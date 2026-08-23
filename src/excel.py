"""The study as a workbook, for the readers who will not open a browser.

Four sheets in the order a reader needs them: what this is and what it does
not claim; the criteria and where each weight comes from; the ranking with
every figure behind it; and the wage gap with all sixteen origins priced so
the comparison can be re-based in the sheet rather than in the tool.

The dashboard cannot hand a file over — its published form runs under a
policy that blocks downloads a page starts itself — so the workbook is built
here and shipped as a file.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src import config as C
from src.baselines import load as baseline_load
from src.dashboard import (
    LIMITS, PILLAR_LABELS, PILLAR_NOTES, PILLARS, SOURCES, payload,
)
# Imported plainly, and this module's own entry point is named build_workbook:
# calling it build shadowed the import, so _ranking re-entered this module and
# recursed until it was killed.
from src.panel import build, with_centres
from src.score import normalise, raw_pillars, rank, score
from src.stability import run as stability_run

OUT = "gbs-location-selection.xlsx"

INK = "1A1A1A"
RULE = "BFBFBF"
HEAD_BG = "14312A"
BAND_BG = "EFEFEA"
SANS = "Calibri"

H1 = Font(name=SANS, size=16, bold=True, color=INK)
H2 = Font(name=SANS, size=11, bold=True, color=INK)
EYEBROW = Font(name=SANS, size=8, bold=True, color="6E7B74")
BODY = Font(name=SANS, size=10, color=INK)
NOTE = Font(name=SANS, size=8, italic=True, color="6E7B74")
TH = Font(name=SANS, size=9, bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
RIGHT = Alignment(horizontal="right", vertical="top")
UNDER = Border(bottom=Side(style="thin", color=RULE))


def _sheet(wb, title, heading, standfirst):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = title
    ws.sheet_view.showGridLines = False
    ws["A1"] = "GBS & GCC LOCATION SELECTION"
    ws["A1"].font = EYEBROW
    ws["A2"] = heading
    ws["A2"].font = H1
    ws["A3"] = standfirst
    ws["A3"].font = Font(name=SANS, size=9.5, color="4D554F")
    ws["A3"].alignment = WRAP
    ws.row_dimensions[3].height = 28
    return ws


def _header(ws, row, labels, widths):
    for i, (label, width) in enumerate(zip(labels, widths), start=1):
        c = ws.cell(row=row, column=i, value=label)
        c.font, c.alignment = TH, WRAP
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _row(ws, row, values, formats=None, banded=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = BODY
        c.alignment = RIGHT if isinstance(v, (int, float)) else WRAP
        c.border = UNDER
        if formats and formats.get(i):
            c.number_format = formats[i]
        if banded:
            c.fill = PatternFill("solid", fgColor=BAND_BG)


def _notes(ws, row, lines):
    for i, line in enumerate(lines):
        c = ws.cell(row=row + i, column=1, value=line)
        c.font = NOTE
        c.alignment = WRAP
    return row + len(lines)


def build_workbook(path: str = OUT) -> str:
    data = payload()
    wb = Workbook()

    _read_me(wb, data)
    _criteria(wb, data)
    _ranking(wb, data)
    _wage_gap(wb, data)
    _not_ranked(wb, data)

    wb.save(path)
    return path


def _read_me(wb, data):
    ws = _sheet(
        wb, "Read me", "What this is",
        "A ranking of cities where GBS and GCC finance roles are genuinely "
        "advertised, scored on seven pillars of public data and re-ranked across "
        "2,000 defensible weightings. Cities the evidence cannot separate share "
        "a band rather than a rank.",
    )
    ws.column_dimensions["A"].width = 104

    row = 5
    for heading, body in (
        ("What it answers", "Which cities survive a change in what you are buying, and which "
                            "differences are too small for the evidence to carry."),
        ("What it does not answer", "Whether a city suits your mandate. Nothing here is a "
                                    "recommendation, and no figure is a client benchmark."),
        ("Population", "GBS and GCC roles only. The broad finance-operations sample was 87% "
                       "retained finance; the filter is what makes Kraków rather than Warsaw "
                       "Poland's leader."),
        ("Precision", "The classifier was audited five times over a hundred postings; two of "
                      "the last twenty were wrong. That error rate is modelled in the stability "
                      "column, not merely noted."),
    ):
        ws.cell(row=row, column=1, value=heading).font = H2
        c = ws.cell(row=row + 1, column=1, value=body)
        c.font, c.alignment = BODY, WRAP
        ws.row_dimensions[row + 1].height = 28
        row += 3

    ws.cell(row=row, column=1, value="Limitations").font = H2
    row += 1
    row = _notes(ws, row, [f"•  {x}" for x in LIMITS])
    row += 1
    ws.cell(row=row, column=1, value=f"Postings sample: {data['asOf']}.  "
            "Method and code: github.com/morichtereur/gbs-location-selection").font = NOTE


def _criteria(wb, data):
    ws = _sheet(
        wb, "Criteria & weights", "The seven criteria",
        "Each weight is that criterion's share of the decision, and the shares "
        "total 100%. The two centre types start from different weightings "
        "because they buy different things.",
    )
    keys = list(C.ARCHETYPES)
    _header(
        ws, 5,
        ["Criterion", "Favours", "What is measured", "Source", "Vintage"]
        + [f"{C.ARCHETYPES[k]['label']} weight" for k in keys],
        [17, 34, 34, 32, 12, 15, 15],
    )
    src = {s["pillar"]: s for s in SOURCES}
    for i, p in enumerate(PILLARS):
        s = src.get(PILLAR_LABELS[p], {})
        _row(ws, 6 + i, [
            PILLAR_LABELS[p], PILLAR_NOTES[p], s.get("detail", ""),
            s.get("name", ""), s.get("vintage", ""),
            *[C.ARCHETYPES[k]["weights"][p] for k in keys],
        ], formats={6: "0%", 7: "0%"}, banded=i % 2 == 1)

    row = 6 + len(PILLARS)
    _row(ws, row, ["Total", "", "", "", "",
                   *[sum(C.ARCHETYPES[k]["weights"].values()) for k in keys]],
         formats={6: "0%", 7: "0%"})
    for i in (1, 6, 7):
        ws.cell(row=row, column=i).font = H2

    row += 2
    for k in keys:
        a = C.ARCHETYPES[k]
        ws.cell(row=row, column=1, value=a["label"]).font = H2
        c = ws.cell(row=row + 1, column=1, value=f"{a['blurb']}  {a['why']}")
        c.font, c.alignment = BODY, WRAP
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=7)
        ws.row_dimensions[row + 1].height = 30
        row += 3

    _notes(ws, row, [
        "Weights are a starting position, not a finding. The dashboard re-ranks across 2,000 "
        "draws around them and reports what survives.",
        "Removing a criterion entirely tests whether it decides anything: the transactional "
        "hub's top band survives losing any pillar except cost; the judgment centre needs five "
        "of seven.",
    ])


def _ranking(wb, data):
    ws = _sheet(
        wb, "City ranking", "The ranking, with every figure behind it",
        "Scores are relative to the cities listed, not absolute ratings. Cities "
        "sharing a band could not be separated by the draws, and their order "
        "inside it carries no information.",
    )
    keys = list(C.ARCHETYPES)
    _header(
        ws, 5,
        ["Centre type", "Rank", "Band", "City", "Country", "Top-3 across reweightings",
         "Postings", "Employers", "Processing share", "Cost USD / month", "Cost basis",
         "Languages", "Operators already there"],
        [16, 6, 6, 15, 14, 13, 9, 10, 12, 13, 21, 20, 46],
    )
    # Ranked here rather than read off the payload: the payload carries the
    # inputs in source order, and a sheet headed "the ranking" that is not one
    # is worse than no sheet.
    panel = {
        k: m for k, m in with_centres(build()).items() if m.complete and m.is_city
    }
    by_id = {r["id"]: r for rows in data["views"]["city"].values() for r in rows}

    row = 6
    for k in keys:
        weights = C.ARCHETYPES[k]["weights"]
        order = rank(score(normalise(raw_pillars(panel, k)), weights))
        st = stability_run(panel, k, weights=weights)
        for i, city in enumerate(order):
            r = by_id[city]
            _row(ws, row, [
                C.ARCHETYPES[k]["label"], i + 1, st.band.get(city),
                r["name"], data["marketNames"].get(r["parent"], r["parent"]),
                st.frequency.get(city), r["postings"], r["employers"],
                r["mixTransactional"], r["cost"],
                (f"city, {r['regionIndex']:.2f}x national"
                 if r.get("regionIndex") else "national"),
                ", ".join(r.get("languages") or []) or "-",
                ", ".join(r.get("operators") or []) or "-",
            ], formats={6: "0%", 9: "0%", 10: "#,##0"}, banded=i % 2 == 1)
            row += 1

    _notes(ws, row + 1, [
        f"Top-3 is the share of {C.DRAWS:,} reweightings in which the city finished in the "
        "top three. Band groups cities the draws could not separate.",
        "Operators exclude staffing firms and merge one company's several spellings. "
        "Languages are those the postings ask for. Neither is scored.",
    ])


def _wage_gap(wb, data):
    ws = _sheet(
        wb, "Wage gap", "What the move is worth, per role per year",
        "The wage line only: no facilities, technology, management overhead, "
        "transition or severance, and headcount held one-for-one. An upper bound "
        "on one component, not a savings case.",
    )
    cities = data["views"]["city"][list(C.ARCHETYPES)[0]]
    bases = data["baselines"]

    _header(ws, 5, ["City", "Country", "Cost USD / month"]
            + [b["label"] for b in bases],
            [15, 14, 14] + [13] * len(bases))
    for i, r in enumerate(sorted(cities, key=lambda r: r["cost"] or 0)):
        _row(ws, 6 + i, [
            r["name"], data["marketNames"].get(r["parent"], r["parent"]), r["cost"],
            *[(b["monthly"] - r["cost"]) * 12 for b in bases],
        ], formats={j: "#,##0" for j in range(3, 4 + len(bases))}, banded=i % 2 == 1)

    row = 6 + len(cities) + 1
    ws.cell(row=row, column=1, value="Origin wage, USD / month").font = H2
    row += 1
    _header(ws, row, ["Origin", "USD / month", "Observed", "Also a candidate?"],
            [22, 14, 11, 17])
    for i, b in enumerate(bases):
        _row(ws, row + 1 + i,
             [b["label"], b["monthly"], b["year"], "yes" if b["scored"] else "no"],
             formats={2: "#,##0"}, banded=i % 2 == 1)

    _notes(ws, row + len(bases) + 2, [
        "A positive figure is the annual wage the move takes out per role; a negative one "
        "means the city is dearer than the origin.",
        "The origin is always a national figure while the Polish cities carry a regional index, "
        "so a capital-city premium sits on one side of the subtraction and not the other.",
        "Germany's earnings come from EU-SILC where the others come from labour force surveys, "
        "and Germany (2022) and Singapore (2021) are the oldest observations in the set.",
    ])


def _not_ranked(wb, data):
    """The dashboard answers "what about Manila"; the workbook has to as well.

    Without this sheet the two deliverables disagree about what the study
    covers, and the workbook is the one that gets forwarded.
    """
    ws = _sheet(
        wb, "Not ranked", "Locations outside the ranking",
        "Five pillars reach these markets because ILOSTAT and the World Bank "
        "cover every country alike. Capability and employer depth do not, and "
        "those are the two the job postings carry, so nothing here is scored "
        "or ranked against the cities that are.",
    )
    _header(
        ws, 5,
        ["City", "Market", "Cost USD / month", "Observed", "Relevant workforce",
         "Governance", "Hours shared with HQ"],
        [18, 20, 15, 10, 17, 12, 18],
    )
    for i, r in enumerate(data["beyond"]):
        _row(ws, 6 + i, [
            r["city"], r["market"], r["cost"], r["costYear"],
            r["talent"], r["risk"], r["overlap"],
        ], formats={3: "#,##0", 4: "0", 5: "#,##0", 6: "0", 7: "0.0"},
             banded=i % 2 == 1)

    row = 6 + len(data["beyond"]) + 1
    ws.cell(row=row, column=1, value="Three other kinds of absence").font = H2
    row += 1
    near = ", ".join(
        f"{m['name']} ({m['postings']} postings, {m['employers']} employers)"
        for m in data["nearMisses"][:4]
    )
    un = " and ".join(f"{c} ({cities})" for c, cities in data["unpriceable"].items())
    row = _notes(ws, row, [
        f"Seen but too thin: {data['nearMissTotal']} locations appear in the sample and clear "
        f"neither threshold. Closest are {near}. The employer count is usually what stops "
        "them, and one employer hiring is an office rather than a centre.",
        f"Not priceable: {un}. ILOSTAT publishes no earnings by occupation for either, so "
        "there is no cost figure on the basis every market here uses.",
        "Incoherent: Egypt reports professionals at 1.06x clerical pay against 1.3-2.9x in "
        "every other market, so it is excluded by test rather than by judgement.",
    ])
    for r_ in range(row - 3, row):
        ws.row_dimensions[r_].height = 30
        ws.merge_cells(start_row=r_, start_column=1, end_row=r_, end_column=7)


def main() -> None:
    print(f"wrote {build_workbook()}")


if __name__ == "__main__":
    main()
