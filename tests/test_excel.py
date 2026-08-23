"""The workbook is the deliverable that leaves the building, so it is tested."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from src import config as C
from src import population
from src.excel import build_workbook

# Both snapshots are gitignored, so CI has no data to build a workbook from.
# importorskip found the duckdb module and let these run anyway, and the build
# failed on the missing file instead of skipping. Two databases are in play:
# this repo's GBS/GCC sample, and the sibling repo's broad one behind the
# contaminant shares. The workbook needs both.
_DBS = (population.DB_PATH, C.POSTINGS_DB)
needs_postings = pytest.mark.skipif(
    not all(db.exists() for db in _DBS),
    reason="postings snapshot missing: "
           + ", ".join(str(db) for db in _DBS if not db.exists()),
)


@pytest.fixture(scope="module")
def book(tmp_path_factory):
    path = tmp_path_factory.mktemp("xl") / "book.xlsx"
    build_workbook(str(path))
    return load_workbook(path)


def test_the_module_entry_point_does_not_shadow_the_panel_builder():
    """This module imports panel.build and once also defined its own build().

    The import lost, so the sheet builder called this module's entry point,
    which called the sheet builder, until the process was killed. It presented
    as a hang with no traceback, so the guard is a name check rather than a
    timeout.
    """
    import src.excel as excel

    assert excel.build.__module__ == "src.panel", (
        f"src.excel.build resolves to {excel.build.__module__}, shadowing the panel builder"
    )


@needs_postings
def test_every_sheet_is_present_and_carries_rows(book):
    assert book.sheetnames == [
        "Read me", "Criteria & weights", "City ranking", "Wage gap", "Not ranked",
    ]
    for name in book.sheetnames:
        assert book[name].max_row > 5, name


@needs_postings
def test_weights_on_the_criteria_sheet_total_one_per_centre_type(book):
    """The sheet states the shares total 100%; a reader will add the column up."""
    ws = book["Criteria & weights"]
    pillars = len(C.ARCHETYPES[list(C.ARCHETYPES)[0]]["weights"])
    for col in (6, 7):
        shares = [ws.cell(row=6 + i, column=col).value for i in range(pillars)]
        assert all(s is not None for s in shares)
        assert abs(sum(shares) - 1.0) < 1e-9, (col, sum(shares))


@needs_postings
def test_the_ranking_sheet_is_actually_ranked(book):
    """It was not, at first: it carried the payload's source order under a title
    that promised a ranking."""
    ws = book["City ranking"]
    seen: dict[str, list[int]] = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row[0] or not isinstance(row[1], int):
            continue
        seen.setdefault(row[0], []).append(row[1])
    assert len(seen) == len(C.ARCHETYPES)
    for label, ranks in seen.items():
        assert ranks == list(range(1, len(ranks) + 1)), (label, ranks)
